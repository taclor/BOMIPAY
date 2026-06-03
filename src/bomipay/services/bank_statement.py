import csv
import hashlib
import io
import json
import logging
import uuid
from datetime import datetime, timezone, timedelta
from typing import Optional

from sqlalchemy import select, and_
from sqlalchemy.ext.asyncio import AsyncSession

from ..models.bank_statement import (
    BankStatementEntry,
    BankStatementImport,
    BankStatementImportStatus,
    BankStatementReconciliation,
    BankStatementReconciliationStatus,
)
from ..models.transaction import Transaction

logger = logging.getLogger("bomipay")

# Try to import openpyxl for XLSX support
try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


def _normalize_hash(merchant_id: str, row: dict) -> str:
    payload = json.dumps(
        {
            "merchant_id": merchant_id,
            "entry_date": str(row.get("entry_date", "")),
            "description": str(row.get("description", "")),
            "debit_amount_minor": row.get("debit_amount_minor", 0),
            "credit_amount_minor": row.get("credit_amount_minor", 0),
            "reference": str(row.get("reference", "")),
            "currency": str(row.get("currency", "NGN")),
        },
        sort_keys=True,
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def parse_csv_rows(file_content: bytes) -> list[dict]:
    reader = csv.DictReader(io.StringIO(file_content.decode("utf-8", errors="replace")))
    rows = []
    for row in reader:
        rows.append(dict(row))
    return rows


def parse_xlsx_rows(file_content: bytes) -> list[dict]:
    """Parse XLSX file and extract rows as dictionaries."""
    if not XLSX_AVAILABLE:
        raise RuntimeError("openpyxl is required for XLSX parsing. Install with: pip install openpyxl")
    
    workbook = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    sheet = workbook.active
    
    # Get headers from first row
    headers = []
    first_row_values = []
    for cell in sheet.iter_rows(min_row=1, max_row=1, values_only=False):
        for c in cell:
            header = str(c.value or "").strip().lower()
            headers.append(header if header else f"col_{len(headers)}")
            first_row_values.append(c.value)
    
    rows = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = {}
        for col_idx, value in enumerate(row):
            if col_idx < len(headers):
                row_dict[headers[col_idx]] = value
        if row_dict and any(v for v in row_dict.values()):
            rows.append(row_dict)
    
    return rows


def normalize_row(raw: dict, currency: str = "NGN") -> Optional[dict]:
    try:
        # Handle case-insensitive field lookups
        def get_field(keys, default=""):
            for k in keys:
                if k in raw and raw[k] is not None:
                    return raw[k]
                # Try lowercase version
                k_lower = k.lower()
                for rk in raw.keys():
                    if rk.lower() == k_lower and raw[rk] is not None:
                        return raw[rk]
            return default

        description = str(get_field(["description", "narration", "details"], "")).strip()
        if not description:
            return None

        def parse_amount(key_list) -> int:
            for k in key_list:
                val = str(get_field([k], "")).replace(",", "").strip()
                if val and val not in ("", "None"):
                    try:
                        return int(float(val) * 100)
                    except (ValueError, TypeError):
                        pass
            return 0

        debit = parse_amount(["debit", "debit_amount", "withdrawal"])
        credit = parse_amount(["credit", "credit_amount", "deposit"])

        entry_date_raw = str(get_field(["date", "entry_date", "transaction_date"], "")).strip()
        if not entry_date_raw or entry_date_raw == "None":
            return None
            
        try:
            # Try ISO format first (with or without time)
            try:
                entry_date = datetime.fromisoformat(entry_date_raw)
            except (ValueError, TypeError):
                # Try basic YYYY-MM-DD format
                try:
                    entry_date = datetime.strptime(entry_date_raw, "%Y-%m-%d")
                except (ValueError, TypeError):
                    # Try YYYY-MM-DD HH:MM:SS format
                    entry_date = datetime.strptime(entry_date_raw, "%Y-%m-%d %H:%M:%S")
        except (ValueError, AttributeError, TypeError):
            logger.debug(f"Failed to parse date: {entry_date_raw}")
            return None

        reference = str(get_field(["reference", "ref", "transaction_id"], "")).strip() or None
        currency_val = str(get_field(["currency"], currency)).strip()

        balance_raw = str(get_field(["balance", "balance_after"], "")).strip()
        balance_after = None
        if balance_raw and balance_raw not in ("", "None"):
            try:
                balance_after = int(float(balance_raw.replace(",", "").strip()) * 100)
            except (ValueError, TypeError):
                pass

        return {
            "entry_date": entry_date,
            "value_date": None,
            "description": description,
            "reference": reference,
            "debit_amount_minor": debit,
            "credit_amount_minor": credit,
            "currency": currency_val,
            "balance_after_minor": balance_after,
            "counterparty_name": str(get_field(["counterparty", "payee"], "")).strip() or None,
        }
    except Exception as e:
        logger.debug(f"Exception in normalize_row: {str(e)}")
        return None


class BankStatementService:
    @staticmethod
    async def create_import(
        db: AsyncSession,
        merchant_id: str,
        file_name: str,
        file_type: str,
        bank_account_id: Optional[str] = None,
        data_source_id: Optional[str] = None,
    ) -> BankStatementImport:
        imp = BankStatementImport(
            id=uuid.uuid4(),
            merchant_id=merchant_id,
            bank_account_id=bank_account_id,
            data_source_id=data_source_id,
            file_name=file_name,
            file_type=file_type,
            status=BankStatementImportStatus.uploaded.value,
        )
        db.add(imp)
        await db.flush()
        return imp

    @staticmethod
    async def process_csv_import(
        db: AsyncSession,
        import_record: BankStatementImport,
        file_content: bytes,
        default_currency: str = "NGN",
    ) -> BankStatementImport:
        import_record.status = BankStatementImportStatus.processing.value
        await db.flush()

        raw_rows = parse_csv_rows(file_content)
        import_record.total_rows = len(raw_rows)

        processed = 0
        failed = 0
        errors: list[dict] = []

        for idx, raw in enumerate(raw_rows):
            normalized = normalize_row(raw, default_currency)
            if normalized is None:
                failed += 1
                errors.append({"row": idx + 1, "error": "Failed to normalize row", "raw": raw})
                continue

            normalized_hash = _normalize_hash(str(import_record.merchant_id), normalized)

            # Idempotency check: skip if hash already exists
            existing = await db.execute(
                select(BankStatementEntry).where(BankStatementEntry.normalized_hash == normalized_hash)
            )
            if existing.scalar_one_or_none() is not None:
                processed += 1
                continue

            entry = BankStatementEntry(
                id=uuid.uuid4(),
                merchant_id=import_record.merchant_id,
                import_id=import_record.id,
                bank_account_id=import_record.bank_account_id,
                entry_date=normalized["entry_date"],
                value_date=normalized["value_date"],
                description=normalized["description"],
                reference=normalized["reference"],
                debit_amount_minor=normalized["debit_amount_minor"],
                credit_amount_minor=normalized["credit_amount_minor"],
                currency=normalized["currency"],
                balance_after_minor=normalized["balance_after_minor"],
                counterparty_name=normalized["counterparty_name"],
                raw_row_json=raw,
                normalized_hash=normalized_hash,
                created_at=datetime.now(timezone.utc),
            )
            db.add(entry)
            processed += 1

        import_record.processed_rows = processed
        import_record.failed_rows = failed
        import_record.error_summary = errors or None
        import_record.status = (
            BankStatementImportStatus.completed.value
            if failed == 0
            else (BankStatementImportStatus.failed.value if processed == 0 else BankStatementImportStatus.completed.value)
        )
        import_record.completed_at = datetime.now(timezone.utc)
        await db.flush()
        logger.info(
            "bank_statement.processed",
            extra={
                "import_id": str(import_record.id),
                "total": import_record.total_rows,
                "processed": processed,
                "failed": failed,
            },
        )
        return import_record

    @staticmethod
    async def process_xlsx_import(
        db: AsyncSession,
        import_record: BankStatementImport,
        file_content: bytes,
        default_currency: str = "NGN",
    ) -> BankStatementImport:
        """Process XLSX import (delegates to CSV logic after parsing)."""
        import_record.status = BankStatementImportStatus.processing.value
        await db.flush()

        try:
            raw_rows = parse_xlsx_rows(file_content)
        except Exception as e:
            import_record.status = BankStatementImportStatus.failed.value
            import_record.error_summary = [{"error": f"Failed to parse XLSX: {str(e)}"}]
            await db.flush()
            logger.error(f"bank_statement.xlsx_parse_error: {str(e)}", extra={"import_id": str(import_record.id)})
            return import_record

        import_record.total_rows = len(raw_rows)
        processed = 0
        failed = 0
        errors: list[dict] = []

        for idx, raw in enumerate(raw_rows):
            normalized = normalize_row(raw, default_currency)
            if normalized is None:
                failed += 1
                errors.append({"row": idx + 1, "error": "Failed to normalize row", "raw": raw})
                continue

            normalized_hash = _normalize_hash(str(import_record.merchant_id), normalized)

            # Idempotency check
            existing = await db.execute(
                select(BankStatementEntry).where(BankStatementEntry.normalized_hash == normalized_hash)
            )
            if existing.scalar_one_or_none() is not None:
                processed += 1
                continue

            entry = BankStatementEntry(
                id=uuid.uuid4(),
                merchant_id=import_record.merchant_id,
                import_id=import_record.id,
                bank_account_id=import_record.bank_account_id,
                entry_date=normalized["entry_date"],
                value_date=normalized["value_date"],
                description=normalized["description"],
                reference=normalized["reference"],
                debit_amount_minor=normalized["debit_amount_minor"],
                credit_amount_minor=normalized["credit_amount_minor"],
                currency=normalized["currency"],
                balance_after_minor=normalized["balance_after_minor"],
                counterparty_name=normalized["counterparty_name"],
                raw_row_json=raw,
                normalized_hash=normalized_hash,
                created_at=datetime.now(timezone.utc),
            )
            db.add(entry)
            processed += 1

        import_record.processed_rows = processed
        import_record.failed_rows = failed
        import_record.error_summary = errors or None
        import_record.status = (
            BankStatementImportStatus.completed.value
            if failed == 0
            else (BankStatementImportStatus.failed.value if processed == 0 else BankStatementImportStatus.completed.value)
        )
        import_record.completed_at = datetime.now(timezone.utc)
        await db.flush()
        logger.info(
            "bank_statement.xlsx_processed",
            extra={
                "import_id": str(import_record.id),
                "total": import_record.total_rows,
                "processed": processed,
                "failed": failed,
            },
        )
        return import_record

    @staticmethod
    async def get_import(db: AsyncSession, import_id: str) -> Optional[BankStatementImport]:
        result = await db.execute(select(BankStatementImport).where(BankStatementImport.id == import_id))
        return result.scalar_one_or_none()

    @staticmethod
    async def list_imports(db: AsyncSession, merchant_id: str) -> list[BankStatementImport]:
        result = await db.execute(
            select(BankStatementImport)
            .where(BankStatementImport.merchant_id == merchant_id)
            .order_by(BankStatementImport.created_at.desc())
        )
        return list(result.scalars().all())

    @staticmethod
    async def list_entries_for_import(
        db: AsyncSession,
        import_id: str,
        merchant_id: str,
        skip: int = 0,
        limit: int = 100,
    ) -> list[BankStatementEntry]:
        result = await db.execute(
            select(BankStatementEntry)
            .where(
                BankStatementEntry.import_id == import_id,
                BankStatementEntry.merchant_id == merchant_id,
            )
            .order_by(BankStatementEntry.entry_date)
            .offset(skip)
            .limit(limit)
        )
        return list(result.scalars().all())

    @staticmethod
    async def list_entries(
        db: AsyncSession,
        merchant_id: str,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        skip: int = 0,
        limit: int = 100,
    ) -> list[BankStatementEntry]:
        stmt = select(BankStatementEntry).where(BankStatementEntry.merchant_id == merchant_id)
        if date_from:
            stmt = stmt.where(BankStatementEntry.entry_date >= date_from)
        if date_to:
            stmt = stmt.where(BankStatementEntry.entry_date <= date_to)
        stmt = stmt.order_by(BankStatementEntry.entry_date.desc()).offset(skip).limit(limit)
        result = await db.execute(stmt)
        return list(result.scalars().all())

    @staticmethod
    async def reconcile_entry(
        db: AsyncSession,
        entry_id: str,
        entry_merchant_id: str,
        entry_currency: str,
        entry_amount: Optional[int],
        entry_reference: Optional[str],
        entry_date,
        import_record_id: str,
    ) -> Optional[BankStatementReconciliation]:
        """Reconcile a single entry by finding matching transactions."""
        # Build query for matching transactions - select only ID to avoid hydrating full objects
        stmt = select(Transaction.id).where(
            and_(
                Transaction.merchant_id == entry_merchant_id,
                Transaction.currency == entry_currency,
            )
        )
        
        # Match on reference if available
        if entry_reference:
            stmt = stmt.where(
                (Transaction.internal_reference == entry_reference)
                | (Transaction.external_reference == entry_reference)
                | (Transaction.provider_transaction_id == entry_reference)
            )
        else:
            # Match on amount and date window (within 7 days)
            if entry_amount:
                stmt = stmt.where(
                    Transaction.amount == entry_amount,
                    Transaction.initiated_at >= entry_date - timedelta(days=7),
                    Transaction.initiated_at <= entry_date + timedelta(days=7),
                )
        
        result = await db.execute(stmt.limit(1))
        transaction_id = result.scalar_one_or_none()
        
        if transaction_id:
            recon = BankStatementReconciliation(
                id=uuid.uuid4(),
                merchant_id=entry_merchant_id,
                import_id=import_record_id,
                entry_id=entry_id,
                transaction_id=transaction_id,
                match_status=BankStatementReconciliationStatus.matched.value,
                confidence_score_bps=9500 if entry_reference else 7500,
                match_notes=f"Matched via {'reference' if entry_reference else 'amount & date'}",
            )
        else:
            recon = BankStatementReconciliation(
                id=uuid.uuid4(),
                merchant_id=entry_merchant_id,
                import_id=import_record_id,
                entry_id=entry_id,
                transaction_id=None,
                match_status=BankStatementReconciliationStatus.unmatched.value,
                confidence_score_bps=0,
                match_notes="No matching transaction found",
            )
        
        db.add(recon)
        return recon

    @staticmethod
    async def reconcile_import(
        db: AsyncSession,
        import_id: str,
        merchant_id: str,
    ) -> dict:
        """Reconcile all entries in an import against transactions."""
        imp = await BankStatementService.get_import(db, import_id)
        if not imp:
            raise ValueError(f"Import {import_id} not found")
        
        if str(imp.merchant_id) != str(merchant_id):
            raise ValueError("Merchant ID mismatch")
        
        # Fetch entry data as tuples (avoid full object hydration to prevent session issues)
        entries_result = await db.execute(
            select(
                BankStatementEntry.id,
                BankStatementEntry.merchant_id,
                BankStatementEntry.currency,
                BankStatementEntry.credit_amount_minor,
                BankStatementEntry.debit_amount_minor,
                BankStatementEntry.reference,
                BankStatementEntry.entry_date,
            ).where(
                BankStatementEntry.import_id == import_id,
                BankStatementEntry.merchant_id == merchant_id,
            )
        )
        entries = entries_result.all()
        
        matched = 0
        unmatched = 0
        
        # Process each entry without hydrating full objects
        for entry_id, entry_merchant_id, entry_currency, credit, debit, reference, entry_date in entries:
            amount = credit or debit
            recon = await BankStatementService.reconcile_entry(
                db, entry_id, entry_merchant_id, entry_currency, amount, reference, entry_date, import_id
            )
            if recon and recon.transaction_id:
                matched += 1
            else:
                unmatched += 1
        
        await db.flush()
        return {"matched": matched, "unmatched": unmatched, "total": len(entries)}
